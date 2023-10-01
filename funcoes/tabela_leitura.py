import os
from openpyxl import load_workbook
import glob

def exibir_dados_da_tabela(nome_tabela):
    if os.path.exists(nome_tabela):
        wb = load_workbook(nome_tabela)
        sheet = wb['bdArvores']
        print('Exibindo dados da tabela:')
        for row in sheet.iter_rows(values_only=True):
            print(row)
    else:
        print(f'A tabela {nome_tabela} não existe.')

def ler_tabela_excel():
    diretorio = "./Tabelas"  # Diretório atual, você pode alterar se desejar
    arquivos = os.listdir(diretorio)
    
    if not arquivos:
        print("Não há tabelas no diretório.")
        return None  # Retorna None se não houver tabelas
    
    print("Tabelas disponíveis:")
    for i, arquivo in enumerate(arquivos, start=1):
        print(f"{i}. {arquivo}")
    
    escolha = input("\nEscolha o número da tabela que deseja ler (ou pressione Enter para voltar ao menu): ")
    
    if escolha == "":
        return None  # Retorna None se o usuário optar por voltar ao menu
    
    try:
        escolha = int(escolha)
        if 1 <= escolha <= len(arquivos):
            arquivo_escolhido = arquivos[escolha - 1]
            caminho_arquivo = os.path.join(diretorio, arquivo_escolhido)

            wb = load_workbook(caminho_arquivo)
            sheet = wb['bdArvores']
            print(f'\nLendo dados da tabela: {arquivo_escolhido}\n')

            # Imprime cabeçalho da tabela com estilo markdown
            print("+------------+------+------------+")
            print("|   Arvore   |  DAP |   Altura   |")
            print("+------------+------+------------+")

            # Imprime os dados formatados
            for row in sheet.iter_rows(values_only=True):
                print(f"| {str(row[0]):<10} | {str(row[1]):<4} | {str(row[2]):<10} |")

            print("+------------+------+------------+")

            while True:
                voltar_opcao = input("Digite 'T' para voltar para a escolha de tabelas ou 'M' para voltar ao menu: ").upper()
                if voltar_opcao == 'T':
                    return None  # Retorna None para voltar à escolha de tabelas
                elif voltar_opcao == 'M':
                    return 'menu'  # Retorna 'menu' para voltar ao menu principal
                else:
                    print("Opção inválida. Digite 'T' para voltar para a escolha de tabelas ou 'M' para voltar ao menu.")
        else:
            print("Escolha inválida.")
    except ValueError:
        print("Escolha inválida.")




def obter_tabela_mais_recente():
    diretorio = "./Tabelas"  # Diretório atual, você pode alterar se desejar
    
    # Obtenha uma lista de todos os arquivos .xlsx no diretório
    arquivos_xlsx = glob.glob(os.path.join(diretorio, "*.xlsx"))
    
    # Ordene a lista de arquivos por data de modificação (o arquivo mais recente será o último)
    arquivos_xlsx.sort(key=os.path.getmtime)
    
    # Verifique se há pelo menos um arquivo .xlsx no diretório
    if arquivos_xlsx:
        # Pegue o nome do arquivo mais recente
        arquivo_mais_recente = arquivos_xlsx[-1]
        return arquivo_mais_recente
    else:
        return None  # Retorna None se não houver tabelas no diretório