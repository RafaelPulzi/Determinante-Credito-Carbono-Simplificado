import random
from time import sleep
import glob
import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
import os


def criar_tabela():
    diretorio = "./Tabelas"  # Diretório atual, você pode alterar se desejar
    nome_base = "BDinfoArvores"
    extensao = ".xlsx"
    contador = 1

    while True:
        nome_tabela = f"{nome_base}_{contador}{extensao}"
        caminho_tabela = os.path.join(diretorio, nome_tabela)
        if not os.path.exists(caminho_tabela):
            tabela = Workbook()
            bdarvores_page = tabela.create_sheet('bdArvores')
            bdarvores_page.append(['arvore', 'dap', 'altura'])  # Adicione os cabeçalhos à nova tabela
            tabela.save(filename=caminho_tabela)
            return caminho_tabela
        contador += 1

def exibir_dados_da_tabela(nome_tabela):
    if os.path.exists(nome_tabela):
        wb = load_workbook(nome_tabela)
        sheet = wb['bdArvores']
        print('Exibindo dados da tabela:')
        for row in sheet.iter_rows(values_only=True):
            print(row)
    else:
        print(f'A tabela {nome_tabela} não existe.')

def ler_tabela_excel():
    diretorio = "./Tabelas"  # Diretório atual, você pode alterar se desejar
    arquivos = os.listdir(diretorio)
    
    if not arquivos:
        print("Não há tabelas no diretório.")
        return None  # Retorna None se não houver tabelas
    
    print("Tabelas disponíveis:")
    for i, arquivo in enumerate(arquivos, start=1):
        print(f"{i}. {arquivo}")
    
    escolha = input("Escolha o número da tabela que deseja ler (ou pressione Enter para voltar ao menu): ")
    
    if escolha == "":
        return None  # Retorna None se o usuário optar por voltar ao menu
    
    try:
        escolha = int(escolha)
        if 1 <= escolha <= len(arquivos):
            arquivo_escolhido = arquivos[escolha - 1]
            caminho_arquivo = os.path.join(diretorio, arquivo_escolhido)

            wb = load_workbook(caminho_arquivo)
            sheet = wb['bdArvores']
            print(f'Lendo dados da tabela: {arquivo_escolhido}')
            for row in sheet.iter_rows(values_only=True):
                print(row)
            
            while True:
                voltar_opcao = input("Digite 'T' para voltar para a escolha de tabelas ou 'M' para voltar ao menu: ").upper()
                if voltar_opcao == 'T':
                    return None  # Retorna None para voltar à escolha de tabelas
                elif voltar_opcao == 'M':
                    return 'menu'  # Retorna 'menu' para voltar ao menu principal
                else:
                    print("Opção inválida. Digite 'T' para voltar para a escolha de tabelas ou 'M' para voltar ao menu.")
        else:
            print("Escolha inválida.")
    except ValueError:
        print("Escolha inválida.")  


def obter_tabela_mais_recente():
    diretorio = "./Tabelas"  # Diretório atual, você pode alterar se desejar
    
    # Obtenha uma lista de todos os arquivos .xlsx no diretório
    arquivos_xlsx = glob.glob(os.path.join(diretorio, "*.xlsx"))
    
    # Ordene a lista de arquivos por data de modificação (o arquivo mais recente será o último)
    arquivos_xlsx.sort(key=os.path.getmtime)
    
    # Verifique se há pelo menos um arquivo .xlsx no diretório
    if arquivos_xlsx:
        # Pegue o nome do arquivo mais recente
        arquivo_mais_recente = arquivos_xlsx[-1]
        return arquivo_mais_recente
    else:
        return None  # Retorna None se não houver tabelas no diretório

#-----------------------------------------------------------------------------------------------------------------------------------------------------

arvore = arvoresParaFor = 0

response = requests.get("https://api.coinbase.com/v2/prices/MCO2-BRL/spot")
data = response.json()
preco = float(data["data"]["amount"])

dados = {}




while True:
    os.system('clear' if os.name == 'posix' else 'cls')  # Limpa o terminal

    print('''
            ======================================
                EMPREENDER CREDITOS DE CARBONO
            ======================================''')
    print('''
    [1] Quanto ganharia em creditos de carbono fazendo uma floresta (incluindo custos)
    [2] Qual foi sua absorcao total de CO2 + Quanto sua floresta gerou de creditos de carbono
    [3] Exibir dados das tabelas
    [0] Sair do Programa
    ''')
    escolherTratamento = input('Qual opcao voce deseja escolher: ')

    try:
        escolherTratamento = int(escolherTratamento)
        os.system('clear' if os.name == 'posix' else 'cls')
    except ValueError:
        print('Por favor, insira um valor numérico válido.')
        continue

    if escolherTratamento == 1:
        print('''
            ======================================
                  INGRESSAR GANHO DE CREDITOS
            ======================================''')
        while True:
            #os.system('clear' if os.name == 'posix' else 'cls') 
            
            print('''
            DIGITE 0 Caso deseje voltar ao menu

            Escolha a sua especie;
            [1] Jatoba-da-Mata
            [2] Guapuruvu
            [3] Pau-jacaré\n''')

            especie = input('Qual especie de arvore deseja escolher: ')

            


            try:
                especie = int(especie)
            except ValueError:
                print('Por favor, insira um valor numérico válido.')
                continue

            if especie == 0:
                break

            if especie == 1:
                dados = {'especie': 'Jatoba-da-Mata', 'espacoOcupadoM': 6.25, 'absorveCO2/ano': 0.12, 'precoSemente': 3, 'crecimentoMaximo/anos': 10}
            elif especie == 2:
                dados = {'especie': 'Guapuruvu', 'espacoOcupadoM': 2.10, 'absorveCO2/ano': 0.04, 'precoSemente': 0.75, 'crecimentoMaximo/anos': 21}
            elif especie == 3:
                 dados = {'especie': 'Pau-jacaré', 'espacoOcupadoM': 2.60, 'absorveCO2/ano': 0.04, 'precoSemente': 1.6, 'crecimentoMaximo/anos': 15}
            else:
                print('[ERRO] NUMERO DIGITADO E INVALIDO, TENTE NOVAMENTE\n')
                continue

            area = input('Qual e a area que voce tem disponivel para a plantacao (m²): ')
            print("\n")
            try:
                area = float(area)
                if area <= 0:
                    raise ValueError("A Area mencionada não pode ser um valor negativo")
            except ValueError:
                print("\n   Por favor, insira um valor numérico válido.\n")
                continue

            N_Arvores = area/dados['espacoOcupadoM']
            gastos = N_Arvores*dados['precoSemente']
            tempoDeAbsorcao = dados['absorveCO2/ano']*N_Arvores*dados['crecimentoMaximo/anos']

            # Dados
            numero_arvores = int(N_Arvores)
            gastos = gastos
            anos_crescimento = dados["crecimentoMaximo/anos"]
            tempo_de_absorcao = tempoDeAbsorcao
            geracao_creditos_carbono = tempo_de_absorcao * preco

            # Criação da tabela
            header = ["Descrição", "Valor"]
            data = [
                ["Número de árvores plantadas", f"{numero_arvores} Árvores"],
                ["Total de gastos", f"R${gastos:.2f}"],
                ["Tempo de absorção", f"{anos_crescimento} anos"],
                ["Absorção de CO2", f"{tempo_de_absorcao:.2f} T/CO2"],
                ["Geração de créditos de carbono", f"R${geracao_creditos_carbono:.2f}"]
            ]

            # Encontre o comprimento máximo de cada coluna
            max_lengths = [max(len(str(item)) for item in column) for column in zip(*data)]

            # Imprime a tabela formatada
            separator = "+".join("-" * (length + 2) for length in max_lengths)
            print(separator)
            print("| " + " | ".join(f"{header[i]:<{max_lengths[i]}}" for i in range(len(header))) + " |")
            print(separator)
            for row in data:
                print("| " + " | ".join(f"{row[i]:<{max_lengths[i]}}" for i in range(len(row))) + " |")
            print(separator)

    elif escolherTratamento == 2:
        print('''
            ======================================
                MEDICAO DE CARBONO NAS ARVORES
            ======================================''')

        print('''
        Todos os valores digitados seram interpretados em uma tabela em excel, assim sendo, toda vez que um looping terminar os dados serao enviados para a tabela, assim que 
        finalizar o programa as tabelas junto de suas interpretacoes seram liberadas para download\n''')


        while True:
            # os.system('clear' if os.name == 'posix' else 'cls') 
            
            arvoresDoLaco = input('Quantas arvores temos dessa especie: ')
            tabela_nome = criar_tabela()
            tabela_existente = load_workbook(tabela_nome)
            bdarvores_page = tabela_existente['bdArvores']

            tabela_nome_mais_recente = obter_tabela_mais_recente()

            try:
                arvoresDoLaco = int(arvoresDoLaco)
                if arvoresDoLaco <= 0:
                    raise ValueError("A Area mencionada não pode ser um valor negativo")
            except ValueError:
                print('Por favor, insira um valor numérico válido.')
                continue

            try:
                dapmin = float(input('Digite o DAP minimo da arvore (m): '))
                dapmax = float(input('Digite o DAP maximo da arvore (m): '))
                if dapmin > dapmax:
                    print('O DAP minimo digitada é maior do que o DAP maximo\n')
                    continue
            except ValueError:
                print('Por favor, insira valores numéricos válidos para o DAP.')
                continue
            
            
            print('--------------------------------------------------------------------------')

            try:
                hmin = float(input('Digite a altura mínima da árvore (m): '))
                hmax = float(input('Digite a altura máxima da árvore (m): '))
                if hmin > hmax:
                    print('A altura minima digitada é maior do que a altura maxima\n')
                    continue
            except ValueError:
                print('Por favor, insira valores numéricos válidos para a altura.')
                continue
            

            print('--------------------------------------------------------------------------')

            arvoresParaFor += arvoresDoLaco
            

            for AdicionandoTabela in range(arvoresParaFor - arvoresDoLaco, arvoresParaFor):
                h = random.randint(hmin, hmax)
                dap = random.randint(dapmin, dapmax)
                arvore += 1

                

                # Adicione os dados à tabela
                bdarvores_page.append([arvore, dap, h])
                
            # Salve a planilha com os novos dados
            tabela_existente.save(filename=tabela_nome)
            print("Dados adicionados à tabela atual.")
                

            opcao = str(input('Deseja continuar[S/N]: ')).upper()
            if opcao == 'N':
                print("\n")
                tabela = pd.read_excel(tabela_nome_mais_recente, sheet_name="bdArvores")

                tabela['C02Evitado'] = (0.013840*(tabela['dap']**2.437632))*(tabela['altura']*0.428609) #Biomassa presa nos fustes em t ha
                tabela['C02Evitado'] = tabela['C02Evitado']*0.5 #EC
                tabela['C02Evitado'] = tabela['C02Evitado']*3.67 #Conversao para Co2 em t ha
                creditosDeCarbono = (tabela['C02Evitado'].sum()).tolist()



                # Dados
                total_carbono_evitado = tabela["C02Evitado"].sum()
                creditos_de_carbono = creditosDeCarbono * preco

                # Criação da tabela
                header = ["Descrição", "Valor"]
                data = [
                    ["CARBONO EVITADO (T / ha)", str(total_carbono_evitado)],
                    ["CREDITOS DE CARBONO", f"R${creditos_de_carbono:.2f}"]
                ]

                # Encontre o comprimento máximo de cada coluna
                max_lengths = [max(len(str(item)) for item in column) for column in zip(*data)]

                # Imprime a tabela formatada
                separator = "+".join("-" * (length + 2) for length in max_lengths)
                print(separator)
                print("| " + " | ".join(f"{header[i]:<{max_lengths[i]}}" for i in range(len(header))) + " |")
                print(separator)
                for row in data:
                    print("| " + " | ".join(f"{row[i]:<{max_lengths[i]}}" for i in range(len(row))) + " |")
                print(separator)
                
                voltarMenu = int(input('''
                [1] VOLTAR PARA O MENU
                [2] CRIAR OUTRA TABELA
                
                Qual sua escolha: '''))
                if voltarMenu == 1:
                    break
                elif voltarMenu == 2:
                    print('Aguarde um instante, estamos reformulando....')
                    sleep(5)

    elif escolherTratamento == 3:
        resultado = ler_tabela_excel()
        if resultado == 'menu':
            break  # Sai do loop principal e encerra o programa
        elif resultado is None:
            continue  # Volta ao início do loop principal
        
    elif escolherTratamento == 0:
        print('Obrigado, volte sempre ;)')
        break

    else:
        print('--[ERRO] NUMERO DIGITADO E INVALIDO, TENTE NOVAMENTE--\n')        